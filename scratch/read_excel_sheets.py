import openpyxl

wb = openpyxl.load_workbook("Combined 401k Prospecting Plan.xlsx", read_only=True)
print("Sheets:", wb.sheetnames)
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows: {len(rows)}")
    if rows:
        print("Header/First row:", rows[0])
        for r in rows[1:4]:
            print("  ", r)
wb.close()
