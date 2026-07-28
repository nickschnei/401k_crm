import os
import sys
import sqlite3
import re
import openpyxl

# Add the project root to sys.path to allow importing modules
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if base_dir not in sys.path:
    sys.path.append(base_dir)

from utils.pdf_generator import compile_diagnostic_pdf
from api.audits import build_custom_outreach_pitch

STOP_WORDS = {
    "inc", "llc", "corp", "group", "co", "pc", "llp", "ltd", 
    "and", "etc", "solutions", "of", "the", "company", "corporation",
    "incorporated", "association", "dental", "dentistry", "design",
    "management", "wealth", "partners", "consultants", "consulting",
    "engineers", "engineering", "financial", "care", "family", "anesthesia",
    "surgical", "orthodontics", "laboratory", "laboratories", "associates",
    "office", "services", "system", "systems", "telecom"
}

def get_significant_words(name):
    if not name:
        return []
    # Lowercase & split by non-alphanumeric
    words = re.findall(r'[a-z0-9]+', name.lower())
    # Filter stop words
    return [w for w in words if w not in STOP_WORDS]

def find_best_sheet_match(p_name, sheet_names):
    p_words = get_significant_words(p_name)
    if not p_words:
        return None
        
    best_sheet = None
    best_score = 0
    
    for s in sheet_names:
        if s == 'Key List':
            continue
        s_words = get_significant_words(s)
        if not s_words:
            continue
            
        intersection = set(p_words) & set(s_words)
        score = len(intersection)
        
        if score > best_score:
            best_score = score
            best_sheet = s
        elif score == best_score and score > 0:
            p_str = "".join(p_words)
            s_str = "".join(s_words)
            if p_str in s_str or s_str in p_str:
                best_sheet = s

    if best_score > 0:
        s_words = get_significant_words(best_sheet)
        min_req = min(len(p_words), len(s_words))
        if best_score >= max(1, min_req * 0.5):
            return best_sheet

    return None

def get_zoominfo_contacts(excel_path, sheet_name):
    if not sheet_name:
        return []
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if len(rows) < 2:
        return []
        
    headers = [str(h).strip().lower() for h in rows[0] if h is not None]
    
    # Map headers to indices
    def get_index(names):
        for name in names:
            if name.lower() in headers:
                return headers.index(name.lower())
        return -1
        
    idx_first = get_index(["first name"])
    idx_last = get_index(["last name"])
    idx_title = get_index(["job title", "title"])
    idx_phone = get_index(["direct phone number", "direct phone", "company hq phone", "phone"])
    idx_email = get_index(["email address", "email", "email_address"])
    idx_linkedin = get_index(["linkedin contact profile url", "linkedin url", "linkedin"])
    idx_city = get_index(["person city", "city"])
    idx_state = get_index(["person state", "state"])

    contacts = []
    for r in rows[1:]:
        if not any(r):
            continue
        first = r[idx_first] if (idx_first != -1 and idx_first < len(r)) else ""
        last = r[idx_last] if (idx_last != -1 and idx_last < len(r)) else ""
        name = f"{first} {last}".strip() or "Unknown Contact"
        
        title = r[idx_title] if (idx_title != -1 and idx_title < len(r)) else "Executive"
        phone = r[idx_phone] if (idx_phone != -1 and idx_phone < len(r)) else "N/A"
        email = r[idx_email] if (idx_email != -1 and idx_email < len(r)) else "N/A"
        linkedin = r[idx_linkedin] if (idx_linkedin != -1 and idx_linkedin < len(r)) else ""
        city = r[idx_city] if (idx_city != -1 and idx_city < len(r)) else ""
        state = r[idx_state] if (idx_state != -1 and idx_state < len(r)) else ""
        location = f"{city}, {state}".strip(", ") or "N/A"
        
        contacts.append({
            "name": name,
            "title": title,
            "phone": str(phone) if phone else "N/A",
            "email": str(email) if email else "N/A",
            "linkedin": str(linkedin) if linkedin else "",
            "location": location
        })
    return contacts

def sanitize_filename(filename: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "", filename).strip()

def main():
    db_path = os.path.join(base_dir, "prospects.db")
    excel_path = os.path.join(base_dir, "Combined 401k Prospecting Plan.xlsx")
    
    if not os.path.exists(db_path):
        print(f"ERROR: Database file not found at {db_path}")
        return
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        return

    # Load sheet names
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    # Define local project workspace target folder
    downloads_dir = os.path.join(base_dir, "Branded_Fiduciary_Diagnostics")
    os.makedirs(downloads_dir, exist_ok=True)
    print(f"Target Output Folder: {downloads_dir}\n")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Query all pipeline prospects
    prospects = cur.execute("SELECT * FROM pipeline_prospects").fetchall()
    print(f"Found {len(prospects)} prospects in pipeline_prospects table.")

    success_count = 0
    for idx, prospect in enumerate(prospects, 1):
        ein = prospect["ein"]
        clean_ein = "".join(c for c in str(ein) if c.isdigit())[-9:].zfill(9)
        employer_name = prospect["employer_name"] or "Prospect"
        
        # Match sheets to retrieve ZoomInfo contacts
        matched_sheet = find_best_sheet_match(employer_name, sheet_names)
        contacts = []
        if matched_sheet:
            try:
                contacts = get_zoominfo_contacts(excel_path, matched_sheet)
                print(f"[{idx}/{len(prospects)}] Matched Sheet '{matched_sheet}' for: {employer_name} (Found {len(contacts)} contacts).")
            except Exception as sheet_err:
                print(f"[{idx}/{len(prospects)}] Error reading sheet '{matched_sheet}': {sheet_err}")
        else:
            print(f"[{idx}/{len(prospects)}] No ZoomInfo sheet matched for: {employer_name}.")

        # Lookup audit records in form_5500_audits
        audit = cur.execute("SELECT * FROM form_5500_audits WHERE ein = ?", (clean_ein,)).fetchone()
        
        if audit:
            record_clean = {
                "ein": clean_ein,
                "employer_name": employer_name,
                "plan_name": audit["plan_name"] or "401(k) Savings Plan",
                "schedule_type": audit["schedule_type"] or "SF",
                "total_assets": float(audit["total_assets"]) if audit["total_assets"] else 0.0,
                "active_participants": audit["active_participants"] or 0,
                "total_eligible_employees": audit["total_eligible_employees"] or 0,
                "admin_expenses": float(audit["admin_expenses"]) if audit["admin_expenses"] else 0.0,
                "corrective_distributions": float(audit["corrective_distributions"]) if audit["corrective_distributions"] else 0.0,
                "participation_rate": float(audit["participation_rate"]) if audit["participation_rate"] else 0.0,
                "fee_ratio": float(audit["fee_ratio"]) if audit["fee_ratio"] else 0.0,
                "compliance_failed": bool(audit["compliance_failed"]),
                "fee_flag": bool(audit["fee_red_flag"]),
                "participation_flag": bool(audit["participation_red_flag"]),
            }
        else:
            record_clean = {
                "ein": clean_ein,
                "employer_name": employer_name,
                "plan_name": "401(k) Savings Plan (Curated Prospect)",
                "schedule_type": "SF",
                "total_assets": float(prospect["total_assets"]) if prospect["total_assets"] else 0.0,
                "active_participants": prospect["active_participants"] or 0,
                "total_eligible_employees": prospect["active_participants"] or 0,
                "admin_expenses": 0.0,
                "corrective_distributions": 0.0,
                "participation_rate": 1.0,
                "fee_ratio": 0.0,
                "compliance_failed": False,
                "fee_flag": False,
                "participation_flag": False,
            }

        try:
            pitch_raw = build_custom_outreach_pitch(record_clean, employer_name)
            # Pass parsed ZoomInfo contacts to compile_diagnostic_pdf
            pdf_stream = compile_diagnostic_pdf(record_clean, pitch_raw, contacts)
            pdf_bytes = pdf_stream.getvalue()

            safe_name = sanitize_filename(f"{employer_name}_{clean_ein}.pdf")
            output_path = os.path.join(downloads_dir, safe_name)

            with open(output_path, "wb") as f:
                f.write(pdf_bytes)
            
            success_count += 1
        except Exception as e:
            print(f"  --> ERROR generating PDF: {e}")

    conn.close()
    print(f"\nCompleted! Generated {success_count} branded diagnostic PDFs successfully.")
    print(f"All reports saved to: {downloads_dir}")

if __name__ == "__main__":
    main()
