from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from api.database import get_async_db
from api.models import Form5500Audit, Prospect
import core
import config
from utils.auth import ClerkUser, get_current_user
from api.prospects import resolve_tenant_id

router = APIRouter()

class AuditResponse(BaseModel):
    ein: str
    schedule_type: Optional[str] = None
    total_assets: Optional[float] = None
    active_participants: Optional[int] = None
    total_eligible_employees: Optional[int] = None
    admin_expenses: Optional[float] = None
    corrective_distributions: Optional[float] = None
    compliance_failed: bool = False
    participation_rate: Optional[float] = None
    fee_ratio: Optional[float] = None
    fee_red_flag: bool = False
    participation_red_flag: bool = False
    found: bool = False

class PitchResponse(BaseModel):
    subject: str
    body: str

def build_custom_outreach_pitch(record: dict | None, employer_name: str) -> str:
    """Programmatically customize the pitch value based on which specific threshold flags are active."""
    name = employer_name if employer_name and str(employer_name).strip() else "your organization"
    
    if not record:
        return f"""Subject: Strategic 401(k) Fiduciary Review for {name}

Dear Plan Sponsor at {name},

I am reaching out to introduce our specialized 401(k) fiduciary advisory services. We assist companies like yours in optimizing plan design, reducing administrative friction, and ensuring fee competitiveness.

We would be glad to run a complimentary plan diagnostic and benchmark report comparing your plan with peer sponsors. This review focuses on:
• Fiduciary oversight and fee validation
• Modernized participant education and auto-enrollment design
• Investment menu optimization and risk management

Are you available for a brief 15-minute introductory conversation next week to see if we can uncover efficiency gains for your plan?

Best regards,
[Your Name]
[Your Fiduciary Advisory Firm]
[Contact Details]"""

    fee_flag = bool(record.get("fee_flag") or record.get("fee_red_flag"))
    part_flag = bool(record.get("participation_flag") or record.get("participation_red_flag"))
    compliance_failed = bool(record.get("compliance_failed"))
    
    fee_ratio = record.get("fee_ratio", 0.0)
    participation_rate = record.get("participation_rate", 0.0)
    fee_bps = int(fee_ratio * 10000)
    
    subject = f"Fiduciary Health & Fee Diagnostic for {name} 401(k) Plan"
    
    lines = [
        f"Subject: {subject}",
        "",
        f"Dear Plan Sponsor at {name},",
        "",
        "I recently conducted a comprehensive fiduciary health audit of the Department of Labor Form 5500 filings "
        f"for the {name} 401(k) plan. Based on our analysis, we identified key opportunities to optimize "
        "your plan design, enhance participant returns, and insulate your committee from regulatory risk.",
        "",
    ]
    
    if fee_flag or part_flag or compliance_failed:
        lines.append("Specifically, our analysis flagged the following areas for immediate review:")
        
        if fee_flag:
            lines.extend([
                f"• **Excessive Plan Fees**: Your plan administrative fee ratio is currently running at {fee_ratio * 100:.2f}% ({fee_bps} bps), which exceeds the 60 basis points industry threshold. As fiduciaries, plan sponsors are legally required to verify that recordkeeping, advisory, and administrative costs are reasonable. We can help execute a fee compression audit to reduce vendor drag and boost employee savings.",
            ])
            
        if part_flag:
            lines.extend([
                f"• **Participation Gap**: The plan's active employee participation rate is {participation_rate * 100:.1f}%, falling below the 70% target. Low participation often triggers discrimination testing limits and restricts executive savings. We specialize in auto-enrollment designs and streamlined education plans to drive engagement.",
            ])
            
        if compliance_failed:
            lines.extend([
                "• **Historic Compliance Alert**: The filing records historic corrective distributions. Compliance testing failures are costly and administratively intensive. We can structure safe harbor provisions or restructure matching parameters to prevent future non-discrimination failures.",
            ])
            
        lines.append("")
    else:
        lines.extend([
            "Your plan's primary indicators currently align well with benchmark thresholds. However, periodic fee validation and fiduciary structural checks are highly recommended to ensure you continue to receive institutional pricing as your assets grow.",
            "",
        ])
        
    lines.extend([
        "We would love to offer your committee a complimentary, side-by-side benchmarking report that details actionable steps to minimize fiduciary risk and elevate employee retirement success.",
        "",
        "Are you open to a brief 15-minute call next Tuesday or Thursday to walk through our findings?",
        "",
        "Best regards,",
        "[Your Name]",
        "[Your Fiduciary Advisory Firm]",
        "[Contact Details]"
    ])
    
    return "\n".join(lines)

@router.get("/{ein}", response_model=AuditResponse)
async def get_fiduciary_audit(
    ein: str, 
    current_user: ClerkUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db)
):
    """Retrieve detailed fiduciary metrics for an EIN, running a chunk-based audit on-demand if missing."""
    try:
        # Resolve user tenant and apply RLS contexts
        tenant_id = await resolve_tenant_id(db, current_user)
        
        # Set session clerk ID for Postgres native RLS policies
        if not db.bind.dialect.name == "sqlite":
            await db.execute(
                text("SELECT set_config('app.current_clerk_id', :clerk_id, true)"),
                {"clerk_id": current_user.clerk_id}
            )
            
        clean_ein = "".join(c for c in str(ein) if c.isdigit())[-9:].zfill(9)
        stmt = select(Form5500Audit).where(Form5500Audit.ein == clean_ein)
        result = await db.execute(stmt)
        record = result.scalar_one_or_none()
        
        # If audit details are missing or hold zero metrics, lazy-trigger the plan audit engine
        if not record or record.total_assets is None or record.total_assets == 0.0:
            try:
                # Release the SQLite transaction lock to prevent deadlocks
                await db.rollback()
                
                from utils.audit_engine import run_plan_audit
                extract_dir = core.ensure_extracted_csvs()
                run_plan_audit(clean_ein, extract_dir)
                
                # Re-query (run_plan_audit already persists findings)
                result = await db.execute(stmt)
                record = result.scalar_one_or_none()
            except Exception as audit_err:
                print(f"Lazy audit failure for EIN {clean_ein}: {audit_err}")

        if not record:
            # Fall back to checking the pipeline_prospects table (tenant-isolated via RLS)
            prospect_stmt = select(Prospect).where(Prospect.ein == clean_ein).where(Prospect.tenant_id == tenant_id)
            prospect_result = await db.execute(prospect_stmt)
            prospect = prospect_result.scalar_one_or_none()
            
            if prospect:
                total_assets = float(prospect.total_assets) if prospect.total_assets is not None else 0.0
                active_participants = prospect.active_participants or 0
                return {
                    "ein": clean_ein,
                    "schedule_type": "Excel",
                    "total_assets": total_assets,
                    "active_participants": active_participants,
                    "total_eligible_employees": active_participants,
                    "admin_expenses": 0.0,
                    "corrective_distributions": 0.0,
                    "compliance_failed": False,
                    "participation_rate": 1.0,
                    "fee_ratio": 0.0,
                    "fee_red_flag": False,
                    "participation_red_flag": False,
                    "found": True
                }
            raise HTTPException(status_code=404, detail=f"No DOL filing or Excel prospect details found for EIN {clean_ein}.")

        return {
            "ein": clean_ein,
            "schedule_type": record.schedule_type,
            "total_assets": float(record.total_assets) if record.total_assets is not None else 0.0,
            "active_participants": record.active_participants,
            "total_eligible_employees": record.total_eligible_employees,
            "admin_expenses": float(record.admin_expenses) if record.admin_expenses is not None else 0.0,
            "corrective_distributions": float(record.corrective_distributions) if record.corrective_distributions is not None else 0.0,
            "compliance_failed": bool(record.compliance_failed),
            "participation_rate": float(record.participation_rate) if record.participation_rate is not None else 0.0,
            "fee_ratio": float(record.fee_ratio) if record.fee_ratio is not None else 0.0,
            "fee_red_flag": bool(record.fee_red_flag),
            "participation_red_flag": bool(record.participation_red_flag),
            "found": True
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{ein}/pitch", response_model=PitchResponse)
async def generate_outreach_pitch(
    ein: str, 
    employer_name: str = "your prospect",
    current_user: ClerkUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db)
):
    """Compile a Form 5500 data-customized outreach pitch based on the plan fiduciary audit records."""
    try:
        # Resolve user tenant and apply RLS contexts
        tenant_id = await resolve_tenant_id(db, current_user)
        
        # Set session clerk ID for Postgres native RLS policies
        if not db.bind.dialect.name == "sqlite":
            await db.execute(
                text("SELECT set_config('app.current_clerk_id', :clerk_id, true)"),
                {"clerk_id": current_user.clerk_id}
            )
            
        clean_ein = "".join(c for c in str(ein) if c.isdigit())[-9:].zfill(9)
        stmt = select(Form5500Audit).where(Form5500Audit.ein == clean_ein)
        result = await db.execute(stmt)
        record = result.scalar_one_or_none()
        
        pitch_dict = None
        if record:
            pitch_dict = {
                "fee_flag": record.fee_red_flag,
                "participation_flag": record.participation_red_flag,
                "compliance_failed": record.compliance_failed,
                "fee_ratio": float(record.fee_ratio) if record.fee_ratio else 0.0,
                "participation_rate": float(record.participation_rate) if record.participation_rate else 0.0,
            }
        else:
            prospect_stmt = select(Prospect).where(Prospect.ein == clean_ein).where(Prospect.tenant_id == tenant_id)
            prospect_result = await db.execute(prospect_stmt)
            prospect = prospect_result.scalar_one_or_none()
            if prospect:
                pitch_dict = {
                    "fee_flag": False,
                    "participation_flag": False,
                    "compliance_failed": False,
                    "fee_ratio": 0.0,
                    "participation_rate": 1.0,
                }
            
        pitch_text = build_custom_outreach_pitch(pitch_dict, employer_name)
        
        # Split subject line from body
        subject = "Fiduciary Health & Fee Diagnostic"
        body = pitch_text
        if pitch_text.startswith("Subject:"):
            parts = pitch_text.split("\n", 1)
            subject = parts[0].replace("Subject:", "").strip()
            body = parts[1].strip() if len(parts) > 1 else ""

        return {
            "subject": subject,
            "body": body
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{ein}/pdf")
async def download_fiduciary_diagnostic_pdf(
    ein: str, 
    current_user: ClerkUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db)
):
    """Compile and stream a branded, highly polished 1-page PDF diagnostic report for a plan audit."""
    try:
        clean_ein = "".join(c for c in str(ein) if c.isdigit())[-9:].zfill(9)
        tenant_id = "default_tenant"
        
        try:
            tenant_id = await resolve_tenant_id(db, current_user)
            if not db.bind.dialect.name == "sqlite":
                await db.execute(
                    text("SELECT set_config('app.current_clerk_id', :clerk_id, true)"),
                    {"clerk_id": current_user.clerk_id}
                )
        except Exception as auth_err:
            print(f"[PDFAuth] Tenant resolution notice: {auth_err}")
            
        stmt = select(Form5500Audit).where(Form5500Audit.ein == clean_ein)
        record = None
        try:
            result = await db.execute(stmt)
            record = result.scalar_one_or_none()
        except Exception as db_err:
            print(f"[PDFDb] Audit lookup notice for EIN {clean_ein}: {db_err}")
        
        # If missing or holding zero metrics, try lazy audit
        if not record or record.total_assets is None or record.total_assets == 0.0:
            try:
                await db.rollback()
                from utils.audit_engine import run_plan_audit
                extract_dir = core.ensure_extracted_csvs()
                run_plan_audit(clean_ein, extract_dir)
                
                result = await db.execute(stmt)
                record = result.scalar_one_or_none()
            except Exception as audit_err:
                print(f"Lazy audit failure for EIN {clean_ein}: {audit_err}")

        record_clean = None
        if record:
            employer_name = record.employer_name or "your prospect"
            record_clean = {
                "ein": record.ein,
                "employer_name": employer_name,
                "plan_name": record.plan_name or "401(k) Savings Plan",
                "schedule_type": record.schedule_type or "SF",
                "total_assets": float(record.total_assets) if record.total_assets else 0.0,
                "active_participants": record.active_participants or 0,
                "total_eligible_employees": record.total_eligible_employees or 0,
                "admin_expenses": float(record.admin_expenses) if record.admin_expenses else 0.0,
                "corrective_distributions": float(record.corrective_distributions) if record.corrective_distributions else 0.0,
                "participation_rate": float(record.participation_rate) if record.participation_rate else 0.0,
                "fee_ratio": float(record.fee_ratio) if record.fee_ratio else 0.0,
                "compliance_failed": bool(record.compliance_failed),
                "fee_flag": bool(record.fee_red_flag),
                "participation_flag": bool(record.participation_red_flag),
            }
        else:
            try:
                prospect_stmt = select(Prospect).where(Prospect.ein == clean_ein)
                prospect_result = await db.execute(prospect_stmt)
                prospect = prospect_result.scalar_one_or_none()
                if prospect:
                    employer_name = prospect.employer_name or "your prospect"
                    record_clean = {
                        "ein": clean_ein,
                        "employer_name": employer_name,
                        "plan_name": "401(k) Savings Plan (Curated Prospect)",
                        "schedule_type": "Excel",
                        "total_assets": float(prospect.total_assets) if prospect.total_assets else 0.0,
                        "active_participants": prospect.active_participants or 0,
                        "total_eligible_employees": prospect.active_participants or 0,
                        "admin_expenses": 0.0,
                        "corrective_distributions": 0.0,
                        "participation_rate": 1.0,
                        "fee_ratio": 0.0,
                        "compliance_failed": False,
                        "fee_flag": False,
                        "participation_flag": False,
                    }
            except Exception as prospect_err:
                print(f"[PDFProspect] Prospect query error for EIN {clean_ein}: {prospect_err}")

        # Final fallback guarantee
        if not record_clean:
            record_clean = {
                "ein": clean_ein,
                "employer_name": f"Organization (EIN {clean_ein})",
                "plan_name": "401(k) Savings Plan",
                "schedule_type": "SF",
                "total_assets": 0.0,
                "active_participants": 0,
                "total_eligible_employees": 0,
                "admin_expenses": 0.0,
                "corrective_distributions": 0.0,
                "participation_rate": 0.0,
                "fee_ratio": 0.0,
                "compliance_failed": False,
                "fee_flag": False,
                "participation_flag": False,
            }

        employer_name = record_clean["employer_name"]
        pitch_raw = build_custom_outreach_pitch(record_clean, employer_name)
        
        contacts = _load_contacts_for_employer(employer_name)
        from utils.pdf_generator import compile_diagnostic_pdf
        from fastapi.responses import StreamingResponse
        
        pdf_stream = compile_diagnostic_pdf(record_clean, pitch_raw, contacts)
        
        filename = f"fiduciary_diagnostic_{clean_ein}.pdf"
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        return StreamingResponse(pdf_stream, media_type="application/pdf", headers=headers)
    except Exception as e:
        print(f"[PDFError] Global error in download_fiduciary_diagnostic_pdf for EIN {ein}: {e}")
        from utils.pdf_generator import compile_diagnostic_pdf
        from fastapi.responses import StreamingResponse
        fallback_record = {
            "ein": clean_ein,
            "employer_name": "Plan Organization",
            "plan_name": "401(k) Savings Plan",
            "schedule_type": "SF",
            "total_assets": 0.0,
            "active_participants": 0,
            "total_eligible_employees": 0,
            "admin_expenses": 0.0,
            "corrective_distributions": 0.0,
            "participation_rate": 0.0,
            "fee_ratio": 0.0,
            "compliance_failed": False,
            "fee_flag": False,
            "participation_flag": False,
        }
        pdf_stream = compile_diagnostic_pdf(fallback_record, "Fiduciary assessment pending.")
        return StreamingResponse(pdf_stream, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=fiduciary_diagnostic_{clean_ein}.pdf"})


@router.get("/{ein}/pdf/short")
async def download_fiduciary_short_pdf(
    ein: str, 
    current_user: ClerkUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Serves the authentic original filed Form 5500 PDF document from the local PDF store if present,
    or redirects to the official Department of Labor EFAST2 portal.
    """
    import os
    import config
    from fastapi.responses import FileResponse, RedirectResponse
    
    clean_ein = "".join(c for c in str(ein) if c.isdigit())[-9:].zfill(9)
    base_dir = str(getattr(config, "BASE_DIR", os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    
    # Check potential local store locations for the real original PDF
    pdf_candidates = [
        os.path.join(base_dir, "data", "pdfs", f"{clean_ein}.pdf"),
        os.path.join(base_dir, "data", "pdfs", f"Form_5500_{clean_ein}.pdf"),
        os.path.join(base_dir, "extracted_data", "pdfs", f"{clean_ein}.pdf"),
        os.path.join(base_dir, "extracted_data", f"{clean_ein}.pdf"),
    ]
    
    for pdf_path in pdf_candidates:
        try:
            if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                print(f"[PDFStore] Serving real original Form 5500 PDF for EIN {clean_ein} from {pdf_path}")
                return FileResponse(
                    path=pdf_path,
                    media_type="application/pdf",
                    filename=f"Form_5500_SF_{clean_ein}.pdf"
                )
        except Exception as check_err:
            print(f"[PDFCandidate] Candidate check error for {pdf_path}: {check_err}")
            
    # Fallback to official DOL EFAST2 filing search portal
    return RedirectResponse(
        url="https://www.efast.dol.gov/5500search/",
        status_code=307
    )


@router.get("/batch/zip")
async def download_batch_branded_pdfs(
    eins: str,
    current_user: ClerkUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Generate and package branded Fiduciary Diagnostic PDFs for multiple plans into a single ZIP archive.
    """
    import io
    import zipfile
    from fastapi.responses import StreamingResponse
    from utils.pdf_generator import compile_diagnostic_pdf
    
    ein_list = [e.strip() for e in eins.split(",") if e.strip()]
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for ein in ein_list:
            clean_ein = "".join(c for c in str(ein) if c.isdigit())[-9:].zfill(9)
            
            stmt = select(Form5500Audit).where(Form5500Audit.ein == clean_ein)
            record = None
            try:
                result = await db.execute(stmt)
                record = result.scalar_one_or_none()
            except Exception as db_err:
                print(f"[BatchZip] DB error for {clean_ein}: {db_err}")
                
            if not record or record.total_assets is None or record.total_assets == 0.0:
                try:
                    await db.rollback()
                    from utils.audit_engine import run_plan_audit
                    extract_dir = core.ensure_extracted_csvs()
                    run_plan_audit(clean_ein, extract_dir)
                    result = await db.execute(stmt)
                    record = result.scalar_one_or_none()
                except Exception as audit_err:
                    print(f"[BatchZip] Lazy audit notice for {clean_ein}: {audit_err}")
                    
            record_clean = None
            if record:
                employer_name = record.employer_name or f"Employer {clean_ein}"
                record_clean = {
                    "ein": record.ein,
                    "employer_name": employer_name,
                    "plan_name": record.plan_name or "401(k) Savings Plan",
                    "schedule_type": record.schedule_type or "SF",
                    "total_assets": float(record.total_assets) if record.total_assets else 0.0,
                    "active_participants": record.active_participants or 0,
                    "total_eligible_employees": record.total_eligible_employees or 0,
                    "admin_expenses": float(record.admin_expenses) if record.admin_expenses else 0.0,
                    "corrective_distributions": float(record.corrective_distributions) if record.corrective_distributions else 0.0,
                    "participation_rate": float(record.participation_rate) if record.participation_rate else 0.0,
                    "fee_ratio": float(record.fee_ratio) if record.fee_ratio else 0.0,
                    "compliance_failed": bool(record.compliance_failed),
                    "fee_flag": bool(record.fee_red_flag),
                    "participation_flag": bool(record.participation_red_flag),
                }
            else:
                try:
                    prospect_stmt = select(Prospect).where(Prospect.ein == clean_ein)
                    prospect_result = await db.execute(prospect_stmt)
                    prospect = prospect_result.scalar_one_or_none()
                    if prospect:
                        employer_name = prospect.employer_name or f"Prospect {clean_ein}"
                        record_clean = {
                            "ein": clean_ein,
                            "employer_name": employer_name,
                            "plan_name": "401(k) Savings Plan (Curated)",
                            "schedule_type": "Excel",
                            "total_assets": float(prospect.total_assets) if prospect.total_assets else 0.0,
                            "active_participants": prospect.active_participants or 0,
                            "total_eligible_employees": prospect.active_participants or 0,
                            "admin_expenses": 0.0,
                            "corrective_distributions": 0.0,
                            "participation_rate": 1.0,
                            "fee_ratio": 0.0,
                            "compliance_failed": False,
                            "fee_flag": False,
                            "participation_flag": False,
                        }
                except Exception as p_err:
                    print(f"[BatchZip] Prospect query error for {clean_ein}: {p_err}")
                    
            if not record_clean:
                record_clean = {
                    "ein": clean_ein,
                    "employer_name": f"Organization (EIN {clean_ein})",
                    "plan_name": "401(k) Savings Plan",
                    "schedule_type": "SF",
                    "total_assets": 0.0,
                    "active_participants": 0,
                    "total_eligible_employees": 0,
                    "admin_expenses": 0.0,
                    "corrective_distributions": 0.0,
                    "participation_rate": 0.0,
                    "fee_ratio": 0.0,
                    "compliance_failed": False,
                    "fee_flag": False,
                    "participation_flag": False,
                }
                
            employer_name = record_clean["employer_name"]
            pitch_raw = build_custom_outreach_pitch(record_clean, employer_name)
            contacts = _load_contacts_for_employer(employer_name)
            pdf_stream = compile_diagnostic_pdf(record_clean, pitch_raw, contacts)
            
            file_friendly_name = employer_name.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace("\"", "_").replace("<", "_").replace(">", "_").replace("|", "_")
            zip_file.writestr(f"fiduciary_diagnostic_{file_friendly_name}_{clean_ein}.pdf", pdf_stream.getvalue())
            
    zip_buffer.seek(0)
    filename = "fiduciary_diagnostic_reports.zip"
    headers = {
        "Content-Disposition": f"attachment; filename={filename}",
        "Access-Control-Expose-Headers": "Content-Disposition"
    }
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)


# Helper functions for ZoomInfo spreadsheet matching and contact parsing
STOP_WORDS = {
    "inc", "llc", "corp", "group", "co", "pc", "llp", "ltd", 
    "and", "etc", "solutions", "of", "the", "company", "corporation",
    "incorporated", "association", "dental", "dentistry", "design",
    "management", "wealth", "partners", "consultants", "consulting",
    "engineers", "engineering", "financial", "care", "family", "anesthesia",
    "surgical", "orthodontics", "laboratory", "laboratories", "associates",
    "office", "services", "system", "systems", "telecom"
}

def _get_significant_words(name):
    if not name:
        return []
    import re
    words = re.findall(r'[a-z0-9]+', name.lower())
    return [w for w in words if w not in STOP_WORDS]

def _find_best_sheet_match(p_name, sheet_names):
    p_words = _get_significant_words(p_name)
    if not p_words:
        return None
    best_sheet = None
    best_score = 0
    for s in sheet_names:
        if s == 'Key List':
            continue
        s_words = _get_significant_words(s)
        if not s_words:
            continue
        intersection = set(p_words) & set(s_words)
        score = len(intersection)
        if score > best_score:
            best_score = score
            best_sheet = s
        elif score == best_score and score > 0:
            p_str = "".join(p_words)
            s_str = "".join(s_words)
            if p_str in s_str or s_str in p_str:
                best_sheet = s
    if best_score > 0:
        s_words = _get_significant_words(best_sheet)
        min_req = min(len(p_words), len(s_words))
        if best_score >= max(1, min_req * 0.5):
            return best_sheet
    return None

def _get_zoominfo_contacts(excel_path, sheet_name):
    if not sheet_name:
        return []
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h).strip().lower() for h in rows[0] if h is not None]
    
    def get_index(names):
        for name in names:
            if name.lower() in headers:
                return headers.index(name.lower())
        return -1
        
    idx_first = get_index(["first name"])
    idx_last = get_index(["last name"])
    idx_title = get_index(["job title", "title"])
    idx_phone = get_index(["direct phone number", "direct phone", "company hq phone", "phone"])
    idx_email = get_index(["email address", "email", "email_address"])
    idx_linkedin = get_index(["linkedin contact profile url", "linkedin url", "linkedin"])
    idx_city = get_index(["person city", "city"])
    idx_state = get_index(["person state", "state"])

    contacts = []
    for r in rows[1:]:
        if not any(r):
            continue
        first = r[idx_first] if (idx_first != -1 and idx_first < len(r)) else ""
        last = r[idx_last] if (idx_last != -1 and idx_last < len(r)) else ""
        name = f"{first} {last}".strip() or "Unknown Contact"
        title = r[idx_title] if (idx_title != -1 and idx_title < len(r)) else "Executive"
        phone = r[idx_phone] if (idx_phone != -1 and idx_phone < len(r)) else "N/A"
        email = r[idx_email] if (idx_email != -1 and idx_email < len(r)) else "N/A"
        linkedin = r[idx_linkedin] if (idx_linkedin != -1 and idx_linkedin < len(r)) else ""
        city = r[idx_city] if (idx_city != -1 and idx_city < len(r)) else ""
        state = r[idx_state] if (idx_state != -1 and idx_state < len(r)) else ""
        location = f"{city}, {state}".strip(", ") or "N/A"
        
        contacts.append({
            "name": name,
            "title": title,
            "phone": str(phone) if phone else "N/A",
            "email": str(email) if email else "N/A",
            "linkedin": str(linkedin) if linkedin else "",
            "location": location
        })
    return contacts

def _load_contacts_for_employer(employer_name):
    import os
    import config
    import openpyxl
    contacts = []
    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(base_dir, "Combined 401k Prospecting Plan.xlsx")
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            matched_sheet = _find_best_sheet_match(employer_name, sheet_names)
            if matched_sheet:
                contacts = _get_zoominfo_contacts(excel_path, matched_sheet)
    except Exception as e:
        print(f"[ContactsLoader] Warning mapping contacts: {e}")
    return contacts






